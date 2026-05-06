from ch00_py.file_toolbox import count_dirs_files, create_path
from ch19_idea_src.idea2brick import (
    IdeaBook,
    add_spark_num_column,
    create_spark_face_spark_nums,
    get_excel_sheet_tuples,
    get_max_spark_num_from_files,
    get_sheets_with_brick_types,
    get_sheets_with_idea_types,
    get_spark_faces_from_df,
    get_spark_faces_from_files,
    get_validated_i_src_idea_type_sheets,
    ideas_sheets_to_brick_sheets,
)
from openpyxl import Workbook as openpyxl_Workbook
from os.path import join as os_path_join
from pandas import (
    DataFrame,
    ExcelWriter as pandas_ExcelWriter,
    isna as pandas_isna,
    read_excel as pandas_read_excel,
)
from pathlib import Path
from pytest import fixture as pytest_fixture, raises as pytest_raises
from ref.keywords import Ch19Keywords as kw, ExampleStrs as exx


def test_IdeaBook_Exists():
    # ESTABLISH / WHEN
    ideabook = IdeaBook()
    # THEN
    assert not ideabook.ideas
    assert set(ideabook.__dict__.keys()) == {f"{kw.idea}s"}


def test_get_spark_faces_from_df_ReturnsObj_Scenario0_Basic():
    # ESTABLISH
    df = DataFrame({kw.spark_face: ["a", "b", "a", "c"]})
    # WHEN
    result = get_spark_faces_from_df(df)
    # THEN
    assert result == {"a", "b", "c"}


def test_get_spark_faces_from_df_ReturnsObj_Scenario1_excludes_nulls():
    # ESTABLISH
    df = DataFrame({kw.spark_face: ["a", None, "b", float("nan")]})
    # WHEN
    result = get_spark_faces_from_df(df)
    # THEN
    assert result == {"a", "b"}


def test_get_spark_faces_from_df_ReturnsObj_Scenario2_MissingColumnReturnsEmptySet():
    # ESTABLISH
    df = DataFrame({"other_col": [1, 2, 3]})
    # WHEN
    result = get_spark_faces_from_df(df)
    # THEN
    assert result == set()


def test_get_spark_faces_from_files_ReturnsObj_Scenario0_Multiple_files(tmp_path):
    # ESTABLISH
    file1 = tmp_path / "file1.xlsx"
    file2 = tmp_path / "file2.xlsx"

    df1 = DataFrame({kw.spark_face: ["a", "b"]})
    df2 = DataFrame({kw.spark_face: ["b", "c"]})

    df1.to_excel(file1, index=False)
    df2.to_excel(file2, index=False)
    # WHEN
    result = get_spark_faces_from_files(tmp_path)
    # THEN
    assert result == {"a", "b", "c"}


def test_get_spark_faces_from_files_ReturnsObj_Scenario1_Multiple_sheets(tmp_path):
    # ESTABLISH
    file1 = tmp_path / "file1.xlsx"

    df1 = DataFrame({kw.spark_face: ["a"]})
    df2 = DataFrame({kw.spark_face: [exx.sue]})

    with pandas_ExcelWriter(file1) as writer:
        df1.to_excel(writer, sheet_name="Sheet1", index=False)
        df2.to_excel(writer, sheet_name="Sheet2", index=False)
    # WHEN
    result = get_spark_faces_from_files(tmp_path)
    # THEN
    assert result == {"a", exx.sue}


def test_get_spark_faces_from_files_ReturnsObj_Scenario2_IgnoresMissingColumn(tmp_path):
    # ESTABLISH
    file1 = tmp_path / "file1.xlsx"

    df1 = DataFrame({kw.spark_face: ["a"]})
    df2 = DataFrame({"other": [1, 2]})  # no spark_face column

    with pandas_ExcelWriter(file1) as writer:
        df1.to_excel(writer, sheet_name="Sheet1", index=False)
        df2.to_excel(writer, sheet_name="Sheet2", index=False)
    # WHEN
    result = get_spark_faces_from_files(tmp_path)
    # THEN
    assert result == {"a"}


def test_get_max_spark_num_from_files_ReturnsObj_Scenario0_MultipleFiles(tmp_path):
    # ESTABLISH
    file1 = tmp_path / "file1.xlsx"
    file2 = tmp_path / "file2.xlsx"
    df1 = DataFrame({kw.spark_num: [1, 2, 3]})
    df2 = DataFrame({kw.spark_num: [4, 5]})
    df1.to_excel(file1, index=False)
    df2.to_excel(file2, index=False)
    # WHEN
    result = get_max_spark_num_from_files(tmp_path)
    # THEN
    assert result == 5


def test_get_max_spark_num_from_files_ReturnsObj_Scenario1_IgnoresInvalidAndConvertsFloats(
    tmp_path,
):
    # ESTABLISH
    file1 = tmp_path / "file1.xlsx"
    df = DataFrame({kw.spark_num: ["10", "bad", None, 7.9]})  # 7.9 -> 7
    df.to_excel(file1, index=False)
    # WHEN
    result = get_max_spark_num_from_files(tmp_path)
    # THEN
    assert result == 10


def test_get_max_spark_num_from_files_ReturnsObj_Scenario2_MultipleSheetsAndMissingColumn(
    tmp_path,
):
    # ESTABLISH
    file1 = tmp_path / "file1.xlsx"
    df1 = DataFrame({kw.spark_num: [1, 20]})
    df2 = DataFrame({"other": [100, 200]})  # no spark_num
    df3 = DataFrame({kw.spark_num: [15]})
    with pandas_ExcelWriter(file1) as writer:
        df1.to_excel(writer, sheet_name="Sheet1", index=False)
        df2.to_excel(writer, sheet_name="Sheet2", index=False)
        df3.to_excel(writer, sheet_name="Sheet3", index=False)

    # WHEN
    result = get_max_spark_num_from_files(tmp_path)
    # THEN
    assert result == 20


def test_create_spark_face_spark_nums_ReturnsObj_Scenario0_Simple():
    # ESTABLISH
    spark_faces = {exx.sue, exx.bob, exx.yao}
    max_spark_num = 11
    # WHEN
    x_dict = create_spark_face_spark_nums(spark_faces, max_spark_num)
    # THEN
    assert x_dict == {exx.bob: 12, exx.sue: 13, exx.yao: 14}


def test_create_spark_face_spark_nums_ReturnsObj_Scenario1_max_spark_num_IsNone():
    # ESTABLISH
    spark_faces = {exx.sue, exx.bob, exx.yao}
    max_spark_num = None
    # WHEN
    x_dict = create_spark_face_spark_nums(spark_faces, max_spark_num)
    # THEN
    assert x_dict == {exx.bob: 1, exx.sue: 2, exx.yao: 3}


def test_create_spark_face_spark_nums_ReturnsObj_Scenario2_No_max_spark_num():
    # ESTABLISH
    spark_faces = {exx.sue, exx.bob, exx.yao}
    # WHEN
    x_dict = create_spark_face_spark_nums(spark_faces)
    # THEN
    assert x_dict == {exx.bob: 1, exx.sue: 2, exx.yao: 3}


def test_add_spark_num_column_SetsAttr_Scenario0_Add_spark_num_Basic():
    # ESTABLISH
    df = DataFrame({kw.spark_face: ["a", "b", "c"]})
    mapping = {"a": 1, "b": 2, "c": 3}
    # WHEN
    add_spark_num_column(df, mapping)
    # THEN
    assert list(df.columns)[0] == kw.spark_num
    assert df[kw.spark_num].tolist() == [1, 2, 3]


def test_add_spark_num_column_SetsAttr_Scenario1_MissingSparkFaceSets_nan():
    # ESTABLISH
    df = DataFrame({kw.spark_face: ["a", "b", "x"]})
    mapping = {"a": 1, "b": 2}
    # WHEN
    add_spark_num_column(df, mapping)
    # THEN
    assert df[kw.spark_num].tolist()[:2] == [1, 2]
    assert pandas_isna(df[kw.spark_num].iloc[2])


def test_add_spark_num_column_SetsAttr_Scenario0_MutatesOriginalDataframe():
    # ESTABLISH
    df = DataFrame({kw.spark_face: ["a", "b"]})
    mapping = {"a": 1, "b": 2}
    assert kw.spark_num not in df.columns
    # WHEN
    add_spark_num_column(df, mapping)
    # THEN
    assert kw.spark_num in df.columns


@pytest_fixture
def excel_dir(tmp_path):
    """Creates a temporary directory with sample Excel files for testing."""

    # File 1: two sheets
    wb1 = openpyxl_Workbook()
    wb1.active.title = "Alpha"
    wb1.create_sheet("Beta")
    wb1.save(tmp_path / "report.xlsx")

    # File 2: one sheet
    wb2 = openpyxl_Workbook()
    wb2.active.title = "Summary"
    wb2.save(tmp_path / "data.xlsx")

    # Non-Excel file (should be ignored)
    (tmp_path / "notes.txt").write_text("ignore me")

    return tmp_path


def test_get_excel_sheet_tuples_ReturnsObj_Scenario0_AllSheetTuples(excel_dir):
    """All (filename, sheet) pairs across every Excel file are returned."""
    # ESTABLISH / WHEN
    result = get_excel_sheet_tuples(str(excel_dir))
    # THEN
    assert ("data.xlsx", "Summary") in result
    assert ("report.xlsx", "Alpha") in result
    assert ("report.xlsx", "Beta") in result
    assert len(result) == 3


def test_get_excel_sheet_tuples_ReturnsObj_Scenario1_SortedList(excel_dir):
    """Returned list is sorted lexicographically by (filename, sheet_name)."""
    # ESTABLISH / WHEN
    result = get_excel_sheet_tuples(str(excel_dir))
    # THEN
    assert result == sorted(result)


def test_get_excel_sheet_tuples_ReturnsObj_Scenario2_EmptyListForNoExcelFiles(
    tmp_path: Path,
):
    """Returns an empty list when the directory contains no Excel files."""
    # ESTABLISH
    (tmp_path / "readme.md").write_text("nothing here")
    # WHEN
    result = get_excel_sheet_tuples(str(tmp_path))
    # THEN
    assert result == []


def test_get_sheets_with_brick_types_ReturnsObj_Scenario0_MatchingTuples(
    tmp_path: Path,
):  # sourcery skip: extract-duplicate-method
    """Only tuples whose sheet_name contains a bk_string are returned."""
    # ESTABLISH
    ideas_excel_dir = tmp_path / "ideas"
    ideas_excel_dir.mkdir()
    wb1 = openpyxl_Workbook()
    wb1.active.title = "bk00102_Sales"
    wb1.create_sheet("Revenue")
    wb1.create_sheet("Costs_bk00105")
    wb1.save(ideas_excel_dir / "x300reports.xlsx")

    wb2 = openpyxl_Workbook()
    wb2.active.title = "Summary"
    wb2.create_sheet("bk00142_Overview")
    wb2.save(ideas_excel_dir / "report.xlsx")

    # WHEN
    result = get_sheets_with_brick_types(ideas_excel_dir)

    # THEN
    assert ("x300reports.xlsx", "bk00102_Sales") in result
    assert ("x300reports.xlsx", "Costs_bk00105") in result
    assert ("report.xlsx", "bk00142_Overview") in result
    assert ("x300reports.xlsx", "Revenue") not in result
    assert ("report.xlsx", "Summary") not in result


def test_get_sheets_with_idea_types_ReturnsObj_Scenario0_MatchingTuples(
    tmp_path: Path,
):  # sourcery skip: extract-duplicate-method
    """Only tuples whose sheet_name contains a bk_string are returned."""
    # ESTABLISH
    ideas_excel_dir = tmp_path / "ideas"
    ideas_excel_dir.mkdir()
    wb1 = openpyxl_Workbook()
    wb1.active.title = "ii00102_Sales"
    wb1.create_sheet("Revenue")
    wb1.create_sheet("Costs_ii00105")
    wb1.save(ideas_excel_dir / "x300reports.xlsx")

    wb2 = openpyxl_Workbook()
    wb2.active.title = "Summary"
    wb2.create_sheet("ii00142_Overview")
    wb2.save(ideas_excel_dir / "report.xlsx")

    # WHEN
    result = get_sheets_with_idea_types(ideas_excel_dir)

    # THEN
    assert ("x300reports.xlsx", "ii00102_Sales") in result
    assert ("x300reports.xlsx", "Costs_ii00105") in result
    assert ("report.xlsx", "ii00142_Overview") in result
    assert ("x300reports.xlsx", "Revenue") not in result
    assert ("report.xlsx", "Summary") not in result


def test_get_validated_i_src_idea_type_sheets_ReturnsObj_Scenario0_IdeaBrSheets(
    tmp_path: Path,
):
    """Returns only brick_type sheet tuples from i_src_dir when there is no overlap."""
    # ESTABLISH
    idea_dir = tmp_path / kw.idea
    idea_dir.mkdir()
    b_src_dir = tmp_path / "bricks"
    b_src_dir.mkdir()
    wb = openpyxl_Workbook()
    wb.active.title = "ii00105_Sales"
    wb.create_sheet("Revenue")
    wb.create_sheet("ii00142_Costs")
    wb.save(idea_dir / "x300reports.xlsx")

    # WHEN
    result = get_validated_i_src_idea_type_sheets(idea_dir, b_src_dir)
    # THEN
    assert ("x300reports.xlsx", "ii00105_Sales") in result
    assert ("x300reports.xlsx", "ii00142_Costs") in result
    assert ("x300reports.xlsx", "Revenue") not in result


# TODO reactive and change test so idea_type sheets are added to bk sheets
# def test_get_validated_i_src_idea_type_sheets_Scenario1_RaisesOnOverlap(
#     tmp_path: Path,
# ):  # sourcery skip: extract-duplicate-method
#     """Raises ValueError when a brick_type sheet name exists in both directories."""
#     # ESTABLISH
#     idea_dir = tmp_path / kw.idea
#     idea_dir.mkdir()
#     idea_wb = openpyxl_Workbook()
#     idea_wb.active.title = "ii00105_Sales"
#     idea_wb.create_sheet("Revenue")
#     idea_wb.create_sheet("ii00142_Costs")
#     x3_filename = "x300reports.xlsx"
#     idea_wb.save(idea_dir / x3_filename)

#     b_src_dir = tmp_path / "brick_overlap"
#     b_src_dir.mkdir()
#     brick_wb = openpyxl_Workbook()
#     brick_wb.active.title = "bk00105_Sales"  # overlaps with idea_dir
#     brick_wb.save(b_src_dir / x3_filename)

#     # WHEN / THEN
#     with pytest_raises(ValueError, match="bk00105_Sales"):
#         get_validated_i_src_idea_type_sheets(idea_dir, b_src_dir)

# TODO figure out if test or function is needed
# def test_get_validated_i_src_idea_type_sheets_Scenario2_DoesNotRaiseError(
#     tmp_path: Path,
# ):  # sourcery skip: extract-duplicate-method
#     """Raises ValueError when a brick_type sheet name exists in both directories."""
#     # ESTABLISH
#     idea_dir = tmp_path / kw.idea
#     idea_dir.mkdir()
#     idea_wb = openpyxl_Workbook()
#     idea_wb.active.title = "ii00105_Sales"
#     idea_wb.create_sheet("Revenue")
#     ii42_sheetname = "ii00142_Costs"
#     idea_wb.create_sheet(ii42_sheetname)
#     x3_filename = "x300reports.xlsx"
#     idea_wb.save(idea_dir / x3_filename)

#     b_src_dir = tmp_path / "brick_overlap"
#     b_src_dir.mkdir()
#     brick_wb = openpyxl_Workbook()
#     brick_wb.active.title = "bk00105_Sales"  # overlaps with idea_dir
#     x4_filename = "x400reports.xlsx"
#     bk42_sheetname = "bk00142_Costs"
#     brick_wb.create_sheet(bk42_sheetname)
#     brick_wb.save(b_src_dir / x4_filename)

#     # WHEN
#     sheet_tuples = get_validated_i_src_idea_type_sheets(idea_dir, b_src_dir)
#     # THEN
#     print(f"{(x3_filename, bk42_sheetname)=}")
#     print(f"{sheet_tuples=}")
#     assert (x3_filename, bk42_sheetname) in sheet_tuples

# # TODO figure out if test or function is needed
# def test_get_validated_i_src_idea_type_sheets_ReturnsObj_Scenario2_EmptyWhenNoIdeaBrSheets(
#     tmp_path: Path,
# ):
#     """Returns an empty list when i_src_dir has no brick_type sheets."""
#     # ESTABLISH
#     b_src_dir = tmp_path / "bricks"
#     b_src_dir.mkdir()
#     wb = openpyxl_Workbook()
#     wb.active.title = "Summary"
#     wb.create_sheet("Details")
#     wb.save(b_src_dir / "brick_report.xlsx")

#     empty_idea = tmp_path / "empty_idea"
#     empty_idea.mkdir()
#     wb = openpyxl_Workbook()
#     wb.active.title = "Summary"
#     wb.save(empty_idea / "plain.xlsx")
#     # WHEN
#     result = get_validated_i_src_idea_type_sheets(empty_idea, b_src_dir)
#     # THEN
#     assert result == []


def test_ideas_sheets_to_brick_sheets_Scenario0_TwoTuples(tmp_path: Path):
    """Returns one (filename, sheet_name) tuple per brick_type sheet copied."""
    # ESTABLISH
    empty_b_src_dir = tmp_path / "bricks"
    empty_b_src_dir.mkdir()

    populated_idea_dir = tmp_path / kw.idea
    populated_idea_dir.mkdir()
    wb = openpyxl_Workbook()
    ws1 = wb.active
    ws1.title = "ii00120_Sales"
    ws1.append(["product", "units", "revenue"])
    ws1.append(["widget", 10, 500])
    ws1.append(["gadget", 5, 250])

    ws2 = wb.create_sheet("ii00104_Costs")
    ws2.append(["category", "amount"])
    ws2.append(["rent", 1000])
    wb.create_sheet("Summary")  # non-BR, should be ignored
    wb.save(populated_idea_dir / "AllSales.xlsx")

    # WHEN
    result = ideas_sheets_to_brick_sheets(populated_idea_dir, empty_b_src_dir)
    # THEN
    dst_all_sales_path = create_path(empty_b_src_dir, "AllSales.xlsx")
    assert (dst_all_sales_path, "bk00104_Costs") in result
    assert (dst_all_sales_path, "bk00120_Sales") in result
    assert len(result) == 2


def test_ideas_sheets_to_brick_sheets_Scenario1_CreatesDestinationFile(
    tmp_path: Path,
):
    """Each copied sheet can be read by pandas and contains the original data."""
    # ESTABLISH
    empty_b_src_dir = tmp_path / "bricks"
    empty_b_src_dir.mkdir()
    populated_idea_dir = tmp_path / kw.idea
    populated_idea_dir.mkdir()
    wb = openpyxl_Workbook()
    ws1 = wb.active
    ws1.title = "ii00120_Sales"
    ws1.append([kw.spark_face, "product", "units", "revenue"])
    ws1.append([exx.sue, "widget", 10, 500])
    ws1.append([exx.sue, "gadget", 5, 250])

    ws2 = wb.create_sheet("ii00104_Costs")
    ws2.append([kw.spark_face, "category", "amount"])
    ws2.append([exx.sue, "rent", 1000])
    wb.create_sheet("Summary")  # non-BR, should be ignored
    wb.save(populated_idea_dir / "AllSales.xlsx")

    # WHEN
    ideas_sheets_to_brick_sheets(populated_idea_dir, empty_b_src_dir)
    # THEN
    allsales_path = os_path_join(str(empty_b_src_dir), "AllSales.xlsx")
    df = pandas_read_excel(allsales_path, sheet_name="bk00120_Sales")
    expected_dst_columns = [kw.spark_num, kw.spark_face, "product", "units", "revenue"]
    assert list(df.columns) == expected_dst_columns
    assert len(df) == 2
    assert df[kw.spark_num].min() == 1
    assert df["revenue"].sum() == 750


# TODO change so sheet data is added
# def test_ideas_sheets_to_brick_sheets_Scenario2_RaisesOnOverlap(tmp_path: Path):
#     # sourcery skip: extract-duplicate-method
#     """Propagates ValueError from get_idea_bk_sheets_validated on sheet name overlap."""
#     # ESTABLISH
#     ideas_dir = tmp_path / kw.idea
#     ideas_dir.mkdir()
#     b_src_dir = tmp_path / "bricks"
#     b_src_dir.mkdir()

#     wb_idea = openpyxl_Workbook()
#     wb_idea.active.title = "bk00120_Sales"
#     allsales_filename = "AllSales.xlsx"
#     wb_idea.save(ideas_dir / allsales_filename)

#     wb_brick = openpyxl_Workbook()
#     wb_brick.active.title = "bk00120_Sales"
#     wb_brick.save(b_src_dir / allsales_filename)
#     # WHEN / THEN
#     with pytest_raises(ValueError, match="bk00120_Sales"):
#         ideas_sheets_to_brick_sheets(ideas_dir, b_src_dir)


def test_ideas_sheets_to_brick_sheets_Scenario3_DestinationFileHas_spark_num_SetBy_b_src_dir(
    tmp_path: Path,
):
    """Each copied sheet can be read by pandas and contains the original data."""
    # ESTABLISH
    b_src_dir = tmp_path / "bricks"
    b_src_dir.mkdir()
    populated_idea_dir = tmp_path / kw.idea
    populated_idea_dir.mkdir()
    brick_wb = openpyxl_Workbook()
    brick_ws1 = brick_wb.active
    brick_ws1.title = "bk00120_Sales"
    expected_dst_columns = [kw.spark_num, kw.spark_face, "product", "units", "revenue"]
    brick_ws1.append(expected_dst_columns)
    curr_spark_num = 10
    brick_ws1.append([curr_spark_num, exx.sue, "widget", 10, 500])
    brick_wb.save(b_src_dir / "OtherFile.xlsx")

    wb = openpyxl_Workbook()
    ws1 = wb.active
    ws1.title = "ii00120_Sales"
    ws1.append([kw.spark_face, "product", "units", "revenue"])
    ws1.append([exx.sue, "widget", 10, 500])
    ws1.append([exx.sue, "gadget", 5, 250])

    ws2 = wb.create_sheet("ii00104_Costs")
    ws2.append([kw.spark_face, "category", "amount"])
    ws2.append([exx.sue, "rent", 1000])
    wb.create_sheet("Summary")  # non-BR, should be ignored
    wb.save(populated_idea_dir / "AllSales.xlsx")

    # WHEN
    ideas_sheets_to_brick_sheets(populated_idea_dir, b_src_dir)
    # THEN
    allsales_path = os_path_join(str(b_src_dir), "AllSales.xlsx")
    df = pandas_read_excel(allsales_path, sheet_name="bk00120_Sales")
    assert list(df.columns) == expected_dst_columns
    assert len(df) == 2
    assert df[kw.spark_num].min() == 11
    assert df[kw.spark_num].min() == curr_spark_num + 1
    assert df["revenue"].sum() == 750


def test_ideas_sheets_to_brick_sheets_Scenario4_ParameterSparkNumAccepted(
    tmp_path: Path,
):
    """Each copied sheet can be read by pandas and contains the original data."""
    # ESTABLISH
    b_src_dir = tmp_path / "bricks"
    b_src_dir.mkdir()
    populated_idea_dir = tmp_path / kw.idea
    populated_idea_dir.mkdir()
    brick_wb = openpyxl_Workbook()
    brick_ws1 = brick_wb.active
    brick_ws1.title = "bk00120_Sales"
    expected_dst_columns = [kw.spark_num, kw.spark_face, "product", "units", "revenue"]
    brick_ws1.append(expected_dst_columns)
    curr_spark_num = 10
    brick_ws1.append([curr_spark_num, exx.sue, "widget", 10, 500])
    brick_wb.save(b_src_dir / "OtherFile.xlsx")

    wb = openpyxl_Workbook()
    ws1 = wb.active
    ws1.title = "ii00120_Sales"
    ws1.append([kw.spark_face, "product", "units", "revenue"])
    ws1.append([exx.sue, "widget", 10, 500])
    ws1.append([exx.sue, "gadget", 5, 250])

    ws2 = wb.create_sheet("ii00104_Costs")
    ws2.append([kw.spark_face, "category", "amount"])
    ws2.append([exx.sue, "rent", 1000])
    wb.create_sheet("Summary")  # non-BR, should be ignored
    wb.save(populated_idea_dir / "AllSales.xlsx")
    db_max_spark_num = 22

    # WHEN
    ideas_sheets_to_brick_sheets(populated_idea_dir, b_src_dir, db_max_spark_num)
    # THEN
    allsales_path = os_path_join(str(b_src_dir), "AllSales.xlsx")
    df = pandas_read_excel(allsales_path, sheet_name="bk00120_Sales")
    assert df[kw.spark_num].min() != 11
    assert df[kw.spark_num].min() != curr_spark_num + 1
    assert df[kw.spark_num].min() == db_max_spark_num + 1


def test_ideas_sheets_to_brick_sheets_Scenario5_src_dir_IsEmptied(
    tmp_path: Path,
):
    """Each copied sheet can be read by pandas and contains the original data."""
    # ESTABLISH
    idea_dir = tmp_path / kw.idea
    idea_dir.mkdir()
    b_src_dir = tmp_path / "bricks"
    b_src_dir.mkdir()

    wb = openpyxl_Workbook()
    ws1 = wb.active
    ws1.title = "bk00120_Sales"
    ws1.append([kw.spark_face, "product", "units", "revenue"])
    ws1.append([exx.sue, "widget", 10, 500])
    wb.save(idea_dir / "AllSales.xlsx")
    assert count_dirs_files(idea_dir) == 1

    # WHEN
    ideas_sheets_to_brick_sheets(idea_dir, b_src_dir)
    # THEN
    assert count_dirs_files(idea_dir) == 0


def test_ideas_sheets_to_brick_sheets_Scenario6_src_num_Exists(
    tmp_path: Path,
):
    """Each copied sheet can be read by pandas and contains the original data."""
    # ESTABLISH
    idea_dir = tmp_path / kw.idea
    idea_dir.mkdir()
    b_src_dir = tmp_path / "bricks"
    b_src_dir.mkdir()

    wb = openpyxl_Workbook()
    ws1 = wb.active
    ws1.title = "ii00120_Sales"
    ws1.append([kw.spark_num, kw.spark_face, "product", "units", "revenue"])
    ws1.append(["", exx.sue, "widget", 10, 500])
    idea_allsales_path = idea_dir / "AllSales.xlsx"
    wb.save(idea_allsales_path)
    assert count_dirs_files(idea_dir) == 1
    assert count_dirs_files(b_src_dir) == 0

    # WHEN
    ideas_sheets_to_brick_sheets(idea_dir, b_src_dir)
    # THEN
    assert count_dirs_files(idea_dir) == 0
    assert count_dirs_files(b_src_dir) == 1
    brick_allsales_path = b_src_dir / "AllSales.xlsx"
    df = pandas_read_excel(brick_allsales_path, sheet_name="bk00120_Sales")
    assert df[kw.spark_num].min() == 1
